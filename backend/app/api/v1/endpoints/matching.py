from typing import Any, List, Dict, Optional
import logging
from fastapi import APIRouter, Depends, HTTPException, status, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc, or_
from sqlalchemy.orm import selectinload
from app.db.session import get_db
from app.core.deps import get_current_active_user, require_recruiter_or_hiring_manager
from app.repositories.job import JobRepository
from app.repositories.resume import ResumeRepository
from app.repositories.match_result import MatchResultRepository
from app.services.matching_engine import get_matching_engine, MatchingRequest
from app.services.embedding_service import EmbeddingService
from app.schemas.matching import (
    MatchRequest,
    MatchResultResponse,
    MatchResultListResponse,
    RankingRequest,
    RankingResponse,
    CandidateRanking,
    MatchStatistics,
    BatchMatchRequest,
    BatchMatchResponse,
    SimilaritySearchRequest,
    SimilaritySearchResponse,
    SimilarityResult,
    StatusUpdateRequest,
    StatusUpdateResponse,
    CandidateStatus
)
from app.models.user import User
from app.models.match_result import MatchResult
import uuid
import time
import asyncio
import csv
import io
from datetime import datetime

router = APIRouter()
logger = logging.getLogger(__name__)


@router.post("/match", response_model=MatchResultListResponse)
async def match_resumes_to_job(
    match_request: MatchRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> Any:
    """
    Match resumes against a job description
    """
    try:
        # Initialize repositories
        job_repo = JobRepository(db)
        resume_repo = ResumeRepository(db)
        match_result_repo = MatchResultRepository(db)
        
        # Verify job exists and user has access
        job = await job_repo.get(match_request.job_id)
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found"
            )
        
        # Check if job is active
        if not job.is_active:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot match against inactive job. Please activate the job first."
            )
        
        # Check if user can access this job
        if job.created_by != current_user.id and current_user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not authorized to match against this job"
            )
        
        # Get matching engine
        matching_engine = await get_matching_engine()
        
        # Create matching request
        engine_request = MatchingRequest(
            job_id=str(match_request.job_id),
            resume_ids=[str(rid) for rid in match_request.resume_ids] if match_request.resume_ids else None,
            include_explanations=match_request.include_explanations,
            min_score_threshold=match_request.min_score_threshold,
            max_results=match_request.max_results
        )
        
        # Execute matching
        matches = await matching_engine.match_resumes_to_job(
            job_repo, resume_repo, match_result_repo, engine_request
        )
        
        # Convert to response format
        match_responses = []
        for match in matches:
            # Get additional data for response
            resume = await resume_repo.get(uuid.UUID(match.resume_id))
            
            match_response = MatchResultResponse(
                id=uuid.uuid4(),  # This would be the actual stored match ID
                job_id=uuid.UUID(match.job_id),
                resume_id=uuid.UUID(match.resume_id),
                match_score=match.overall_score,
                confidence_score=match.confidence,
                explanation=match.explanation,
                key_strengths=match.matched_skills,
                missing_skills=match.missing_skills,
                skill_matches={
                    "matched": match.matched_skills,
                    "missing": match.missing_skills,
                    "additional": match.additional_skills
                },
                created_at=match.created_at,
                updated_at=match.created_at,
                created_by=current_user.id,
                job_title=job.title,
                job_company=job.company,
                candidate_name=resume.candidate_name if resume else "Unknown",
                candidate_email=resume.email if resume else None
            )
            match_responses.append(match_response)
        
        return MatchResultListResponse(
            matches=match_responses,
            total=len(match_responses),
            skip=0,
            limit=len(match_responses)
        )
        
    except Exception as e:
        # Log the full traceback for debugging, then return a generic 500 to client
        logger.exception("Failed to perform matching")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to perform matching: {str(e)}"
        )


@router.get("/results/{job_id}", response_model=MatchResultListResponse)
async def get_match_results_for_job(
    job_id: uuid.UUID,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> Any:
    """
    Get all match results for a job
    """
    try:
        # Initialize repositories
        job_repo = JobRepository(db)
        match_result_repo = MatchResultRepository(db)
        resume_repo = ResumeRepository(db)
        
        # Verify job exists and user has access
        job = await job_repo.get(job_id)
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found"
            )
        
        # Check if user is admin, job creator, or in the same company
        if job.created_by != current_user.id and current_user.role != "admin":
            from app.repositories.user import UserRepository
            user_repo = UserRepository(db)
            job_creator = await user_repo.get(job.created_by)
            
            if not job_creator or job_creator.company_id != current_user.company_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Not authorized to view matches for this job"
                )
        
        # Get match results with resume data - return empty array if none exist
        matches = await match_result_repo.get_by_job(job_id, skip, limit)
        
        # Convert to response format
        match_responses = []
        for match in matches:
            # Get resume data
            resume = await resume_repo.get(match.resume_id)
            
            # Convert resume to dict if it exists
            resume_dict = None
            if resume:
                resume_dict = {
                    "id": str(resume.id),
                    "candidate_name": resume.candidate_name,
                    "email": resume.email,
                    "phone": resume.phone,
                    "content": resume.content[:500] if resume.content else "",  # Truncate for response
                    "uploaded_at": resume.uploaded_at.isoformat() if resume.uploaded_at else None
                }
            
            # Create match response with resume data
            match_response = MatchResultResponse(
                id=match.id,
                job_id=match.job_id,
                resume_id=match.resume_id,
                match_score=float(match.match_score),
                confidence_score=float(match.confidence_score) if match.confidence_score else None,
                explanation=match.explanation,
                key_strengths=match.key_strengths or [],
                missing_skills=match.missing_skills or [],
                skill_matches=match.skill_matches,
                experience_score=float(match.experience_score) if match.experience_score else None,
                skills_score=float(match.skills_score) if match.skills_score else None,
                education_score=float(match.education_score) if match.education_score else None,
                status=match.status if hasattr(match, 'status') and match.status else CandidateStatus.PENDING,
                status_updated_at=match.status_updated_at if hasattr(match, 'status_updated_at') else None,
                status_updated_by=match.status_updated_by if hasattr(match, 'status_updated_by') else None,
                created_at=match.created_at,
                updated_at=match.updated_at,
                created_by=match.created_by,
                job_title=job.title,
                job_company=job.company,
                candidate_name=resume.candidate_name if resume else "Unknown",
                candidate_email=resume.email if resume else None,
                resume=resume_dict
            )
            match_responses.append(match_response)
        
        return MatchResultListResponse(
            matches=match_responses,
            total=len(match_responses),
            skip=skip,
            limit=limit
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get match results: {str(e)}"
        )


@router.get("/job/{job_id}/ranking", response_model=RankingResponse)
async def get_job_candidate_ranking(
    job_id: uuid.UUID,
    min_score: float = Query(0.0, ge=0.0, le=100.0),
    max_results: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> Any:
    """
    Get ranked candidates for a job
    """
    try:
        # Initialize repositories
        job_repo = JobRepository(db)
        match_result_repo = MatchResultRepository(db)
        
        # Verify job exists and user has access
        job = await job_repo.get(job_id)
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found"
            )
        
        # Check if user is admin, job creator, or in the same company
        if job.created_by != current_user.id and current_user.role != "admin":
            from app.repositories.user import UserRepository
            user_repo = UserRepository(db)
            job_creator = await user_repo.get(job.created_by)
            
            if not job_creator or job_creator.company_id != current_user.company_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Not authorized to view rankings for this job"
                )
        
        # Get match results
        matches = await match_result_repo.get_top_matches_for_job(
            job_id, min_score, max_results
        )
        
        # Convert to candidate rankings
        candidates = []
        for i, match in enumerate(matches, 1):
            candidate = CandidateRanking(
                resume_id=match.resume_id,
                candidate_name=match.resume.candidate_name,
                candidate_email=match.resume.email,
                match_score=float(match.match_score),
                confidence_score=float(match.confidence_score) if match.confidence_score else None,
                key_strengths=match.key_strengths or [],
                missing_skills=match.missing_skills or [],
                rank=i
            )
            candidates.append(candidate)
        
        # Calculate statistics
        all_scores = [float(match.match_score) for match in matches]
        avg_score = sum(all_scores) / len(all_scores) if all_scores else 0.0
        top_score = max(all_scores) if all_scores else 0.0
        
        return RankingResponse(
            job_id=job_id,
            job_title=job.title,
            job_company=job.company,
            candidates=candidates,
            total_candidates=len(candidates),
            avg_score=round(avg_score, 2),
            top_score=round(top_score, 2)
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get candidate ranking: {str(e)}"
        )


@router.get("/job/{job_id}/statistics", response_model=MatchStatistics)
async def get_job_match_statistics(
    job_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> Any:
    """
    Get matching statistics for a job
    """
    try:
        # Initialize repositories
        job_repo = JobRepository(db)
        match_result_repo = MatchResultRepository(db)
        
        # Verify job exists and user has access
        job = await job_repo.get(job_id)
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found"
            )
        
        # Check if user is admin, job creator, or in the same company
        if job.created_by != current_user.id and current_user.role != "admin":
            from app.repositories.user import UserRepository
            user_repo = UserRepository(db)
            job_creator = await user_repo.get(job.created_by)
            
            if not job_creator or job_creator.company_id != current_user.company_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Not authorized to view statistics for this job"
                )
        
        # Get all matches for the job
        matches = await match_result_repo.get_by_job(job_id)
        
        if not matches:
            return MatchStatistics(
                job_id=job_id,
                total_candidates=0,
                avg_score=0.0,
                top_score=0.0,
                candidates_above_70=0,
                candidates_above_50=0,
                most_common_missing_skills=[],
                score_distribution={"high": 0, "medium": 0, "low": 0}
            )
        
        # Calculate statistics
        scores = [float(match.match_score) for match in matches]
        avg_score = sum(scores) / len(scores)
        top_score = max(scores)
        
        candidates_above_70 = len([s for s in scores if s >= 70])
        candidates_above_50 = len([s for s in scores if s >= 50])
        
        # Analyze missing skills
        all_missing_skills = []
        for match in matches:
            if match.missing_skills:
                all_missing_skills.extend(match.missing_skills)
        
        # Count skill frequency
        skill_counts = {}
        for skill in all_missing_skills:
            skill_counts[skill] = skill_counts.get(skill, 0) + 1
        
        most_common_missing = sorted(
            skill_counts.items(), 
            key=lambda x: x[1], 
            reverse=True
        )[:10]
        
        # Score distribution
        high_scores = len([s for s in scores if s >= 80])
        medium_scores = len([s for s in scores if 50 <= s < 80])
        low_scores = len([s for s in scores if s < 50])
        
        return MatchStatistics(
            job_id=job_id,
            total_candidates=len(matches),
            avg_score=round(avg_score, 2),
            top_score=round(top_score, 2),
            candidates_above_70=candidates_above_70,
            candidates_above_50=candidates_above_50,
            most_common_missing_skills=[skill for skill, _ in most_common_missing],
            score_distribution={
                "high": high_scores,
                "medium": medium_scores,
                "low": low_scores
            }
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get match statistics: {str(e)}"
        )


@router.get("/resume/{resume_id}/matches", response_model=MatchResultListResponse)
async def get_resume_matches(
    resume_id: uuid.UUID,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
) -> Any:
    """
    Get job matches for a resume
    """
    try:
        # Initialize repositories
        resume_repo = ResumeRepository(db)
        match_result_repo = MatchResultRepository(db)
        
        # Verify resume exists and user has access
        resume = await resume_repo.get(resume_id)
        if not resume:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Resume not found"
            )
        
        # Recruiters and hiring managers can access all resumes
        
        # Get matches
        matches = await match_result_repo.get_by_resume(resume_id, skip, limit)
        
        # Convert to response format
        match_responses = []
        for match in matches:
            match_response = MatchResultResponse(
                id=match.id,
                job_id=match.job_id,
                resume_id=match.resume_id,
                match_score=float(match.match_score),
                confidence_score=float(match.confidence_score) if match.confidence_score else None,
                explanation=match.explanation,
                key_strengths=match.key_strengths or [],
                missing_skills=match.missing_skills or [],
                skill_matches=match.skill_matches,
                experience_score=float(match.experience_score) if match.experience_score else None,
                skills_score=float(match.skills_score) if match.skills_score else None,
                education_score=float(match.education_score) if match.education_score else None,
                status=match.status if hasattr(match, 'status') and match.status else CandidateStatus.PENDING,
                status_updated_at=match.status_updated_at if hasattr(match, 'status_updated_at') else None,
                status_updated_by=match.status_updated_by if hasattr(match, 'status_updated_by') else None,
                created_at=match.created_at,
                updated_at=match.updated_at,
                created_by=match.created_by,
                job_title=match.job.title,
                job_company=match.job.company,
                candidate_name=resume.candidate_name,
                candidate_email=resume.email
            )
            match_responses.append(match_response)
        
        return MatchResultListResponse(
            matches=match_responses,
            total=len(match_responses) + skip,  # Approximation
            skip=skip,
            limit=limit
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get resume matches: {str(e)}"
        )


@router.post("/similarity-search", response_model=SimilaritySearchResponse)
async def similarity_search(
    search_request: SimilaritySearchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
) -> Any:
    """
    Search for similar jobs or resumes using semantic similarity
    """
    try:
        # Generate embedding for query
        embedding_service = EmbeddingService()
        query_embedding = await embedding_service.generate_embedding(search_request.query_text)
        
        results = []
        
        if search_request.search_type == "jobs":
            # Search similar jobs
            job_repo = JobRepository(db)
            similar_jobs = await job_repo.similarity_search(
                query_embedding, 
                search_request.limit, 
                search_request.threshold
            )
            
            for job_data in similar_jobs:
                result = SimilarityResult(
                    id=job_data["id"],
                    title=job_data["title"],
                    content_preview=job_data["description"][:200] + "...",
                    similarity=job_data["similarity"]
                )
                results.append(result)
        
        else:  # search_type == "resumes"
            # Search similar resumes (only for recruiters/hiring managers)
            if current_user.role not in ["recruiter", "hiring_manager", "admin"]:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Not authorized to search resumes"
                )
            
            resume_repo = ResumeRepository(db)
            similar_resumes = await resume_repo.similarity_search(
                query_embedding, 
                search_request.limit, 
                search_request.threshold
            )
            
            for resume_data in similar_resumes:
                result = SimilarityResult(
                    id=resume_data["id"],
                    title=resume_data["candidate_name"],
                    content_preview=resume_data["content"][:200] + "...",
                    similarity=resume_data["similarity"]
                )
                results.append(result)
        
        return SimilaritySearchResponse(
            query=search_request.query_text,
            search_type=search_request.search_type,
            results=results,
            total_found=len(results)
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to perform similarity search: {str(e)}"
        )


@router.get("/match/{match_id}", response_model=MatchResultResponse)
async def get_match_result(
    match_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
) -> Any:
    """
    Get specific match result by ID
    """
    try:
        match_result_repo = MatchResultRepository(db)
        match = await match_result_repo.get(match_id)
        
        if not match:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Match result not found"
            )
        
        # Check access permissions for recruiters and hiring managers
        if current_user.role in ["recruiter", "hiring_manager"]:
            # Recruiters can only see matches for jobs they created
            if match.job.created_by != current_user.id and current_user.role != "admin":
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Not authorized to view this match result"
                )
        
        return MatchResultResponse(
            id=match.id,
            job_id=match.job_id,
            resume_id=match.resume_id,
            match_score=float(match.match_score),
            confidence_score=float(match.confidence_score) if match.confidence_score else None,
            explanation=match.explanation,
            key_strengths=match.key_strengths or [],
            missing_skills=match.missing_skills or [],
            skill_matches=match.skill_matches,
            experience_score=float(match.experience_score) if match.experience_score else None,
            skills_score=float(match.skills_score) if match.skills_score else None,
            education_score=float(match.education_score) if match.education_score else None,
            status=match.status if hasattr(match, 'status') and match.status else CandidateStatus.PENDING,
            status_updated_at=match.status_updated_at if hasattr(match, 'status_updated_at') else None,
            status_updated_by=match.status_updated_by if hasattr(match, 'status_updated_by') else None,
            created_at=match.created_at,
            updated_at=match.updated_at,
            created_by=match.created_by,
            job_title=match.job.title,
            job_company=match.job.company,
            candidate_name=match.resume.candidate_name,
            candidate_email=match.resume.email
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get match result: {str(e)}"
        )



# Status management endpoints
@router.patch("/results/{match_id}/status", response_model=StatusUpdateResponse)
async def update_match_status(
    match_id: uuid.UUID,
    status_update: StatusUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> Any:
    """
    Update the status of a match result (candidate status management)
    """
    match_result_repo = MatchResultRepository(db)
    
    # Get the match result
    match_result = await match_result_repo.get(match_id)
    if not match_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Match result not found"
        )
    
    # Update the status
    from datetime import datetime
    match_result.status = status_update.status
    match_result.status_updated_at = datetime.utcnow()
    match_result.status_updated_by = current_user.id
    
    await db.commit()
    await db.refresh(match_result)
    
    return StatusUpdateResponse(
        id=match_result.id,
        status=match_result.status,
        status_updated_at=match_result.status_updated_at,
        status_updated_by=match_result.status_updated_by,
        message=f"Status updated to {match_result.status.value}"
    )


@router.get("/results/{job_id}/by-status/{status}", response_model=MatchResultListResponse)
async def get_matches_by_status(
    job_id: uuid.UUID,
    status: CandidateStatus,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> Any:
    """
    Get match results filtered by status
    """
    # Build query
    query = select(MatchResult).options(
        selectinload(MatchResult.job),
        selectinload(MatchResult.resume)
    ).where(
        MatchResult.job_id == job_id,
        MatchResult.status == status
    ).order_by(desc(MatchResult.match_score)).offset(skip).limit(limit)
    
    result = await db.execute(query)
    matches = result.scalars().all()
    
    # Get total count
    count_query = select(func.count(MatchResult.id)).where(
        MatchResult.job_id == job_id,
        MatchResult.status == status
    )
    count_result = await db.execute(count_query)
    total = count_result.scalar()
    
    # Format response
    match_responses = []
    for match in matches:
        match_responses.append(MatchResultResponse(
            id=str(match.id),
            job_id=str(match.job_id),
            resume_id=str(match.resume_id),
            match_score=float(match.match_score),
            explanation=match.explanation,
            key_strengths=match.key_strengths or [],
            missing_skills=match.missing_skills or [],
            created_at=match.created_at.isoformat(),
            candidate_name=match.resume.candidate_name if match.resume else None,
            candidate_email=match.resume.email if match.resume else None,
            job_title=match.job.title if match.job else None,
            job_company=match.job.company if match.job else None,
            status=match.status,
            status_updated_at=match.status_updated_at.isoformat() if match.status_updated_at else None,
            status_updated_by=str(match.status_updated_by) if match.status_updated_by else None
        ))
    
    return MatchResultListResponse(
        matches=match_responses,
        total=total,
        skip=skip,
        limit=limit
    )


@router.get("/candidates/all", response_model=MatchResultListResponse)
async def get_all_candidates(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status: Optional[str] = Query(None),
    job_id: Optional[uuid.UUID] = Query(None),
    min_score: Optional[float] = Query(None, ge=0.0, le=100.0),
    max_score: Optional[float] = Query(None, ge=0.0, le=100.0),
    search: Optional[str] = Query(None),
    sort_by: str = Query("created_at", regex="^(created_at|match_score|status)$"),
    sort_order: str = Query("desc", regex="^(asc|desc)$"),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> Any:
    """
    Get all candidates across all jobs with filtering, search, and pagination.
    
    Query Parameters:
    - skip: Number of records to skip (pagination)
    - limit: Maximum number of records to return
    - status: Filter by candidate status (pending, shortlisted, rejected, maybe)
    - job_id: Filter by specific job
    - min_score: Minimum match score filter
    - max_score: Maximum match score filter
    - search: Search by candidate name or email
    - sort_by: Sort field (created_at, match_score, status)
    - sort_order: Sort order (asc, desc)
    - date_from: Filter by date range start
    - date_to: Filter by date range end
    """
    try:
        from sqlalchemy import or_, and_
        
        # Initialize repositories
        job_repo = JobRepository(db)
        match_result_repo = MatchResultRepository(db)
        resume_repo = ResumeRepository(db)
        
        # Build query
        query = select(MatchResult).options(
            selectinload(MatchResult.job),
            selectinload(MatchResult.resume)
        )
        
        # Filter by jobs created by current user (unless admin)
        if current_user.role != "admin":
            query = query.join(MatchResult.job).where(
                MatchResult.job.has(created_by=current_user.id)
            )
        
        # Apply filters
        if job_id:
            query = query.where(MatchResult.job_id == job_id)
        
        # Status filter (if status field exists)
        if status:
            # Check if status column exists in the model
            if hasattr(MatchResult, 'status'):
                query = query.where(MatchResult.status == status)
        
        # Score filters
        if min_score is not None:
            query = query.where(MatchResult.match_score >= min_score)
        if max_score is not None:
            query = query.where(MatchResult.match_score <= max_score)
        
        # Date range filters
        if date_from:
            query = query.where(MatchResult.created_at >= date_from)
        if date_to:
            query = query.where(MatchResult.created_at <= date_to)
        
        # Search by candidate name or email
        if search:
            # Join with resume table for search
            from app.models.resume import Resume
            query = query.join(MatchResult.resume).where(
                or_(
                    Resume.candidate_name.ilike(f"%{search}%"),
                    Resume.email.ilike(f"%{search}%")
                )
            )
        
        # Get total count before pagination
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0
        
        # Apply sorting
        if sort_by == "match_score":
            if sort_order == "desc":
                query = query.order_by(desc(MatchResult.match_score))
            else:
                query = query.order_by(MatchResult.match_score)
        elif sort_by == "status" and hasattr(MatchResult, 'status'):
            if sort_order == "desc":
                query = query.order_by(desc(MatchResult.status))
            else:
                query = query.order_by(MatchResult.status)
        else:  # created_at
            if sort_order == "desc":
                query = query.order_by(desc(MatchResult.created_at))
            else:
                query = query.order_by(MatchResult.created_at)
        
        # Apply pagination
        query = query.offset(skip).limit(limit)
        
        # Execute query
        result = await db.execute(query)
        matches = result.scalars().all()
        
        # Convert to response format
        match_responses = []
        for match in matches:
            # Get resume data
            resume = match.resume
            
            match_response = MatchResultResponse(
                id=match.id,
                job_id=match.job_id,
                resume_id=match.resume_id,
                match_score=float(match.match_score),
                confidence_score=float(match.confidence_score) if match.confidence_score else None,
                explanation=match.explanation,
                key_strengths=match.key_strengths or [],
                missing_skills=match.missing_skills or [],
                skill_matches=match.skill_matches,
                experience_score=float(match.experience_score) if match.experience_score else None,
                skills_score=float(match.skills_score) if match.skills_score else None,
                education_score=float(match.education_score) if match.education_score else None,
                status=match.status if hasattr(match, 'status') and match.status else CandidateStatus.PENDING,
                status_updated_at=match.status_updated_at.isoformat() if hasattr(match, 'status_updated_at') and match.status_updated_at else None,
                status_updated_by=str(match.status_updated_by) if hasattr(match, 'status_updated_by') and match.status_updated_by else None,
                created_at=match.created_at,
                updated_at=match.updated_at,
                created_by=match.created_by,
                job_title=match.job.title if match.job else "Unknown",
                job_company=match.job.company if match.job else "Unknown",
                candidate_name=resume.candidate_name if resume else "Unknown",
                candidate_email=resume.email if resume else None,
                resume={
                    "id": str(resume.id),
                    "candidate_name": resume.candidate_name,
                    "email": resume.email,
                    "phone": resume.phone,
                    "uploaded_at": resume.uploaded_at.isoformat() if resume.uploaded_at else None
                } if resume else None
            )
            match_responses.append(match_response)
        
        return MatchResultListResponse(
            matches=match_responses,
            total=total,
            skip=skip,
            limit=limit
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to get all candidates")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get all candidates: {str(e)}"
        )


@router.get("/jobs/{job_id}/candidates/export")
async def export_candidates(
    job_id: uuid.UUID,
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    status: Optional[str] = Query(None),
    min_score: Optional[float] = Query(None, ge=0.0, le=100.0),
    max_score: Optional[float] = Query(None, ge=0.0, le=100.0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> StreamingResponse:
    """
    Export candidates for a job to CSV or Excel format
    
    Query Parameters:
    - format: Export format (csv or xlsx)
    - status: Filter by candidate status (pending, shortlisted, rejected, maybe)
    - min_score: Minimum match score filter
    - max_score: Maximum match score filter
    """
    try:
        # Initialize repositories
        job_repo = JobRepository(db)
        match_result_repo = MatchResultRepository(db)
        resume_repo = ResumeRepository(db)
        
        # Verify job exists and user has access
        job = await job_repo.get(job_id)
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Job not found"
            )
        
        # Check if user is admin, job creator, or in the same company
        if job.created_by != current_user.id and current_user.role != "admin":
            from app.repositories.user import UserRepository
            user_repo = UserRepository(db)
            job_creator = await user_repo.get(job.created_by)
            
            if not job_creator or job_creator.company_id != current_user.company_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Not authorized to export candidates for this job"
                )
        
        # Get all match results for the job
        matches = await match_result_repo.get_by_job(job_id, skip=0, limit=10000)
        
        # Apply filters
        filtered_matches = []
        for match in matches:
            # Status filter
            if status and hasattr(match, 'status'):
                if match.status != status:
                    continue
            
            # Score filters
            if min_score is not None and match.match_score < min_score:
                continue
            if max_score is not None and match.match_score > max_score:
                continue
            
            filtered_matches.append(match)
        
        # Sort by match score descending
        filtered_matches.sort(key=lambda x: x.match_score, reverse=True)
        
        # Prepare data for export
        export_data = []
        for idx, match in enumerate(filtered_matches, 1):
            # Get resume data
            resume = await resume_repo.get(match.resume_id)
            
            row = {
                "Rank": idx,
                "Candidate Name": resume.candidate_name if resume else "Unknown",
                "Email": resume.email if resume else "",
                "Phone": resume.phone if resume and resume.phone else "",
                "Match Score": f"{match.match_score:.2f}%",
                "Status": getattr(match, 'status', 'pending').title() if hasattr(match, 'status') else "Pending",
                "Key Strengths": ", ".join(match.key_strengths[:5]) if match.key_strengths else "",
                "Missing Skills": ", ".join(match.missing_skills[:5]) if match.missing_skills else "",
                "Applied Date": match.created_at.strftime("%Y-%m-%d %H:%M") if match.created_at else "",
            }
            export_data.append(row)
        
        # Generate export based on format
        if format == "csv":
            # Create CSV
            output = io.StringIO()
            if export_data:
                fieldnames = export_data[0].keys()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(export_data)
            
            # Prepare response
            output.seek(0)
            filename = f"{job.title.replace(' ', '_')}_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                }
            )
        
        elif format == "xlsx":
            # For Excel format, we need openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Candidates"
                
                # Add header with styling
                if export_data:
                    headers = list(export_data[0].keys())
                    ws.append(headers)
                    
                    # Style header row
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Add data rows
                    for row_data in export_data:
                        ws.append(list(row_data.values()))
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                filename = f"{job.title.replace(' ', '_')}_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f"attachment; filename={filename}"
                    }
                )
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="Excel export requires openpyxl package. Please use CSV format instead."
                )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to export candidates")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export candidates: {str(e)}"
        )


@router.get("/candidates/export")
async def export_all_candidates(
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    status: Optional[str] = Query(None),
    job_id: Optional[uuid.UUID] = Query(None),
    min_score: Optional[float] = Query(None, ge=0.0, le=100.0),
    max_score: Optional[float] = Query(None, ge=0.0, le=100.0),
    search: Optional[str] = Query(None),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter_or_hiring_manager)
) -> StreamingResponse:
    """
    Export all candidates across all jobs to CSV or Excel format
    
    Query Parameters:
    - format: Export format (csv or xlsx)
    - status: Filter by candidate status (pending, shortlisted, rejected, maybe)
    - job_id: Filter by specific job
    - min_score: Minimum match score filter
    - max_score: Maximum match score filter
    - search: Search by candidate name or email
    - date_from: Filter by date range start
    - date_to: Filter by date range end
    """
    try:
        from app.models.job import JobDescription
        
        # Initialize repositories
        job_repo = JobRepository(db)
        match_result_repo = MatchResultRepository(db)
        resume_repo = ResumeRepository(db)
        
        # Build query
        query = select(MatchResult).options(
            selectinload(MatchResult.job),
            selectinload(MatchResult.resume)
        )
        
        # Filter by jobs created by current user (unless admin)
        if current_user.role != "admin":
            query = query.join(MatchResult.job).where(
                MatchResult.job.has(created_by=current_user.id)
            )
        
        # Apply filters
        if job_id:
            query = query.where(MatchResult.job_id == job_id)
        
        # Status filter
        if status and hasattr(MatchResult, 'status'):
            query = query.where(MatchResult.status == status)
        
        # Score filters
        if min_score is not None:
            query = query.where(MatchResult.match_score >= min_score)
        if max_score is not None:
            query = query.where(MatchResult.match_score <= max_score)
        
        # Date range filters
        if date_from:
            query = query.where(MatchResult.created_at >= date_from)
        if date_to:
            query = query.where(MatchResult.created_at <= date_to)
        
        # Search by candidate name or email
        if search:
            from app.models.resume import Resume
            query = query.join(MatchResult.resume).where(
                or_(
                    Resume.candidate_name.ilike(f"%{search}%"),
                    Resume.email.ilike(f"%{search}%")
                )
            )
        
        # Sort by match score descending
        query = query.order_by(desc(MatchResult.match_score))
        
        # Execute query (limit to 10000 for export)
        query = query.limit(10000)
        result = await db.execute(query)
        matches = result.scalars().all()
        
        # Prepare data for export
        export_data = []
        for idx, match in enumerate(matches, 1):
            # Get resume data
            resume = match.resume
            job = match.job
            
            row = {
                "Rank": idx,
                "Candidate Name": resume.candidate_name if resume else "Unknown",
                "Email": resume.email if resume else "",
                "Phone": resume.phone if resume and resume.phone else "",
                "Job Title": job.title if job else "Unknown",
                "Company": job.company if job else "Unknown",
                "Match Score": f"{match.match_score:.2f}%",
                "Status": getattr(match, 'status', 'pending').title() if hasattr(match, 'status') else "Pending",
                "Key Strengths": ", ".join(match.key_strengths[:5]) if match.key_strengths else "",
                "Missing Skills": ", ".join(match.missing_skills[:5]) if match.missing_skills else "",
                "Applied Date": match.created_at.strftime("%Y-%m-%d %H:%M") if match.created_at else "",
            }
            export_data.append(row)
        
        # Generate export based on format
        if format == "csv":
            # Create CSV
            output = io.StringIO()
            if export_data:
                fieldnames = export_data[0].keys()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(export_data)
            
            # Prepare response
            output.seek(0)
            filename = f"all_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                }
            )
        
        elif format == "xlsx":
            # For Excel format, we need openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "All Candidates"
                
                # Add header with styling
                if export_data:
                    headers = list(export_data[0].keys())
                    ws.append(headers)
                    
                    # Style header row
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Add data rows
                    for row_data in export_data:
                        ws.append(list(row_data.values()))
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                filename = f"all_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f"attachment; filename={filename}"
                    }
                )
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="Excel export requires openpyxl package. Please use CSV format instead."
                )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to export all candidates")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export all candidates: {str(e)}"
        )
